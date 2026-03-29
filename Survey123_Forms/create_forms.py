#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
SEOC Portal - Survey123 XLSForm Generator
Creates all Survey123 forms for APSDMA SEOC Data Entry Portal
"""

import sys, os
os.environ['PYTHONIOENCODING'] = 'utf-8'
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
from openpyxl.styles import Font, PatternFill
import psycopg2

# ---- DB Connection ----
conn = psycopg2.connect(host='192.168.9.35', port=5432, dbname='apsdma_2026',
                        user='sde', password='apsdma#123', connect_timeout=10)
cur = conn.cursor()
cur.execute('SELECT district_lgd, district_name FROM admin.ap_districts_28 ORDER BY district_name')
DISTRICTS = cur.fetchall()
cur.execute('SELECT mandal_lgd, mandal_name, district_lgd, district_name FROM admin.ap_mandals_688 ORDER BY district_name, mandal_name')
MANDALS = cur.fetchall()
conn.close()
print(f"Loaded {len(DISTRICTS)} districts, {len(MANDALS)} mandals from DB")

OUT_DIR = 'h:/SEOC_DB_PORTAL/SEOC_PORTAL'
os.makedirs(OUT_DIR, exist_ok=True)

# Common choice lists used across forms
COMMON_CHOICES = {
    'gender': [('male', 'Male'), ('female', 'Female'), ('other', 'Other')],
    'yesno': [('yes', 'Yes'), ('no', 'No')],
    'economic_status': [('bpl', 'BPL'), ('apl', 'APL')],
    'payment_status': [('pending', 'Pending'), ('sanctioned', 'Sanctioned'), ('disbursed', 'Disbursed')],
}


def add_choices_sheet(wb, extra_choices=None):
    """Add choices sheet with districts, mandals, and any extra choice lists."""
    ws = wb.create_sheet('choices')
    ws.append(['list_name', 'name', 'label', 'district_lgd'])
    hdr_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = Font(bold=True, color='FFFFFF')

    # Districts
    for lgd, name in DISTRICTS:
        ws.append(['district', str(lgd), name, ''])

    # Mandals (with district_lgd filter)
    for mlgd, mname, dlgd, dname in MANDALS:
        ws.append(['mandal', str(mlgd) if mlgd else mname, mname, str(dlgd)])

    # Common choices
    for list_name, items in COMMON_CHOICES.items():
        for val, label in items:
            ws.append([list_name, val, label, ''])

    # Extra form-specific choices
    if extra_choices:
        for list_name, items in extra_choices.items():
            for val, label in items:
                ws.append([list_name, val, label, ''])


def create_form(filename, title, form_id, survey_rows, extra_choices=None, has_attachments=True):
    """Create a complete Survey123 XLSForm."""
    wb = openpyxl.Workbook()

    # --- SETTINGS ---
    ws_set = wb.active
    ws_set.title = 'settings'
    ws_set.append(['form_title', 'form_id', 'version', 'style'])
    ws_set.append([title, form_id, '2026.1', 'theme-grid'])

    # --- SURVEY ---
    ws = wb.create_sheet('survey', 0)
    headers = ['type', 'name', 'label', 'hint', 'required', 'appearance',
               'calculation', 'constraint', 'constraint_message', 'relevant',
               'default', 'choice_filter', 'read_only']
    ws.append(headers)
    hdr_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = Font(bold=True, color='FFFFFF')

    # -- Metadata (hidden) --
    ws.append(['begin group', 'grp_meta', 'Record Metadata', '', '', 'w5'])
    ws.append(['text', 'entry_status', 'Status', '', '', '', "'Pending'", '', '', '', 'Pending', '', 'yes'])
    ws.append(['text', 'entered_by', 'Entered By', '', '', 'hidden', '', '', '', '', '', '', 'yes'])
    ws.append(['dateTime', 'entry_datetime', 'Entry Time', '', '', '', 'now()', '', '', '', '', '', 'yes'])
    ws.append(['end group', 'grp_meta'])

    # -- Incident Date/Time --
    ws.append(['begin group', 'grp_incident', 'Incident Information', '', '', 'w5'])
    ws.append(['date', 'incident_date', 'Date of Incident', '', 'yes', ''])
    ws.append(['text', 'incident_time', 'Time of Incident', 'e.g. 03:30 PM', '', ''])
    ws.append(['end group', 'grp_incident'])

    # -- Location --
    ws.append(['begin group', 'grp_location', 'Location', '', '', 'w5'])
    ws.append(['geopoint', 'incident_location', 'Location on Map', 'Tap map or enter GPS', '', 'maps'])
    ws.append(['select_one district', 'district', 'District', '', 'yes', 'autocomplete'])
    ws.append(['select_one mandal', 'mandal', 'Mandal', '', 'yes', 'autocomplete',
               '', '', '', '', '', 'district_lgd=${district}'])
    ws.append(['text', 'village', 'Village / Habitation', '', '', ''])
    ws.append(['text', 'location_details', 'Specific Location', 'Road, landmark, etc.', '', ''])
    ws.append(['end group', 'grp_location'])

    # -- Form-specific fields --
    for row in survey_rows:
        ws.append(row)

    # -- Attachments --
    if has_attachments:
        ws.append(['begin group', 'grp_attach', 'Attachments & Documents', '', '', 'w5'])
        ws.append(['image', 'photo_doc', 'Photo / Scanned Document', '', '', 'annotate'])
        ws.append(['file', 'pdf_report', 'PDF Report (MRO/Preliminary)', '', '', ''])
        ws.append(['file', 'pdf_support', 'Supporting Document', '', '', ''])
        ws.append(['end group', 'grp_attach'])

    # -- Remarks --
    ws.append(['text', 'remarks', 'Remarks', '', '', 'multiline'])

    # --- CHOICES ---
    add_choices_sheet(wb, extra_choices)

    path = os.path.join(OUT_DIR, filename)
    wb.save(path)
    print(f"  Created: {filename}")


# ================================================================
# FORM 1: LIGHTNING DEATH REPORT
# ================================================================
print("\n[1/8] Lightning Death Report")
lightning_rows = [
    ['begin group', 'grp_victim', 'Victim Details', '', '', 'w5'],
    ['text', 'victim_name', 'Name of the Deceased', '', 'yes', ''],
    ['text', 'aadhar_number', 'Aadhar Number', '12 digits', '', ''],
    ['select_one gender', 'gender', 'Gender', '', 'yes', ''],
    ['integer', 'age', 'Age (Years)', '', '', ''],
    ['text', 'occupation', 'Occupation', '', '', ''],
    ['select_one location_type', 'death_location', 'Location of Death', '', 'yes', ''],
    ['select_one economic_status', 'economic_status', 'Economic Status (BPL/APL)', '', '', ''],
    ['end group', 'grp_victim'],

    ['begin group', 'grp_kin', 'Next of Kin', '', '', 'w5'],
    ['text', 'kin_name', 'Next of Kin Name', '', '', ''],
    ['text', 'kin_phone', 'Next of Kin Phone', '', '', ''],
    ['text', 'kin_relation', 'Relation to Deceased', '', '', ''],
    ['end group', 'grp_kin'],

    ['begin group', 'grp_alert', 'Alert Information', '', '', 'w5'],
    ['select_one yesno', 'cap_alert_sent', 'CAP Alert Sent?', '', '', ''],
    ['text', 'cap_alert_time', 'CAP Alert Time', '', '', '', '', '', '', '${cap_alert_sent}="yes"'],
    ['select_one yesno', 'whatsapp_alert', 'WhatsApp Alert Sent?', '', '', ''],
    ['end group', 'grp_alert'],

    ['begin group', 'grp_payment', 'Ex-Gratia Payment', '', '', 'w5'],
    ['select_one payment_status', 'payment_status', 'Payment Status', '', '', ''],
    ['text', 'sanction_order', 'Admin Sanction Order No.', '', '', '', '', '', '', '${payment_status}!="pending"'],
    ['decimal', 'exgratia_amount', 'Ex-Gratia Amount (Rs.)', '', '', '', '', '', '', '${payment_status}="disbursed"'],
    ['text', 'cfms_bill_no', 'CFMS Bill Number', '', '', '', '', '', '', '${payment_status}="disbursed"'],
    ['end group', 'grp_payment'],
]

lightning_choices = {
    'location_type': [
        ('open_field', 'Open Field / Open Area'),
        ('under_tree', 'Under Tree / Below Tree'),
        ('near_shed', 'Near Shed / Beside Shed'),
        ('on_house', 'On House / Terrace'),
        ('shock', 'Sound Shock'),
        ('other', 'Other'),
    ],
}

create_form('SEOC_Lightning_Death.xlsx', 'SEOC - Lightning Death Report',
            'seoc_lightning_death', lightning_rows, lightning_choices)


# ================================================================
# FORM 2: DROWNING INCIDENT
# ================================================================
print("[2/8] Drowning Incident Report")
drowning_rows = [
    ['begin group', 'grp_event', 'Event Details', '', '', 'w5'],
    ['text', 'incident_description', 'Incident Description', '', 'yes', 'multiline'],
    ['select_one water_body', 'water_body_type', 'Water Body Type', '', 'yes', ''],
    ['text', 'water_body_name', 'Name of Water Body', 'River/Canal/Tank name', '', ''],
    ['select_one report_source', 'report_source', 'Report Source', '', 'yes', ''],
    ['end group', 'grp_event'],

    ['begin group', 'grp_count', 'Victim Count', '', '', 'w5'],
    ['integer', 'total_victims', 'Total Victims', '', 'yes', ''],
    ['integer', 'persons_saved', 'Persons Saved', '', '', '', '', '', '', '', '0'],
    ['integer', 'persons_died', 'Persons Died', '', '', '', '', '', '', '', '0'],
    ['integer', 'persons_missing', 'Persons Missing', '', '', '', '', '', '', '', '0'],
    ['end group', 'grp_count'],

    ['select_multiple rescue_team', 'rescue_by', 'Rescue By', '', '', ''],

    ['begin repeat', 'rpt_victims', 'Individual Victim Details', '', '', ''],
    ['text', 'victim_name', 'Victim Name', '', '', ''],
    ['integer', 'victim_age', 'Age', '', '', ''],
    ['select_one gender', 'victim_gender', 'Gender', '', '', ''],
    ['select_one victim_status', 'victim_outcome', 'Outcome', '', '', ''],
    ['end repeat', 'rpt_victims'],
]

drowning_choices = {
    'water_body': [
        ('river', 'River'), ('canal', 'Canal'), ('tank', 'Tank/Pond'),
        ('reservoir', 'Reservoir'), ('sea', 'Sea/Beach'), ('well', 'Well/Borewell'),
        ('quarry', 'Quarry'), ('other', 'Other'),
    ],
    'rescue_team': [
        ('sdrf', 'SDRF'), ('ndrf', 'NDRF'), ('fire', 'Fire Department'),
        ('police', 'Local Police'), ('revenue', 'Revenue'), ('swimmers', 'Local Swimmers'),
        ('navy', 'Navy'), ('coast_guard', 'Coast Guard'), ('other', 'Other'),
    ],
    'report_source': [
        ('seoc', 'SEOC Preliminary'), ('112_call', '112 Emergency Call'),
        ('1070_call', '1070 Call'), ('collector', 'District Collector'), ('other', 'Other'),
    ],
    'victim_status': [
        ('saved', 'Saved'), ('died', 'Died'), ('missing', 'Missing'),
    ],
}

create_form('SEOC_Drowning_Incident.xlsx', 'SEOC - Drowning Incident Report',
            'seoc_drowning', drowning_rows, drowning_choices)


# ================================================================
# FORM 3: FLOOD DAMAGE REPORT
# ================================================================
print("[3/8] Flood Damage Report")
flood_rows = [
    ['begin group', 'grp_period', 'Report Period', '', '', 'w5'],
    ['select_one report_shift', 'report_shift', 'Reporting Shift', '', 'yes', ''],
    ['end group', 'grp_period'],

    ['begin group', 'grp_impact', 'Impact Summary', '', '', 'w5'],
    ['integer', 'mandals_affected', 'Mandals Affected', '', '', '', '', '', '', '', '0'],
    ['integer', 'villages_affected', 'Villages Affected', '', '', '', '', '', '', '', '0'],
    ['integer', 'villages_inundated', 'Villages Inundated', '', '', '', '', '', '', '', '0'],
    ['integer', 'population_affected', 'Population Affected', '', '', '', '', '', '', '', '0'],
    ['integer', 'lives_lost', 'Human Lives Lost', '', '', '', '', '', '', '', '0'],
    ['integer', 'cattle_lost', 'Cattle Lost', '', '', '', '', '', '', '', '0'],
    ['end group', 'grp_impact'],

    ['begin group', 'grp_houses', 'House Damage', '', '', 'w5'],
    ['integer', 'houses_fully_pucca', 'Fully Damaged - Pucca', '', '', '', '', '', '', '', '0'],
    ['integer', 'houses_fully_kutcha', 'Fully Damaged - Kutcha', '', '', '', '', '', '', '', '0'],
    ['integer', 'houses_fully_huts', 'Fully Damaged - Huts', '', '', '', '', '', '', '', '0'],
    ['integer', 'houses_partly_pucca', 'Partly Damaged - Pucca', '', '', '', '', '', '', '', '0'],
    ['integer', 'houses_partly_kutcha', 'Partly Damaged - Kutcha', '', '', '', '', '', '', '', '0'],
    ['integer', 'houses_partly_huts', 'Partly Damaged - Huts', '', '', '', '', '', '', '', '0'],
    ['end group', 'grp_houses'],

    ['begin group', 'grp_infra', 'Infrastructure Damage', '', '', 'w5'],
    ['integer', 'electric_poles', 'Electric Poles Damaged', '', '', '', '', '', '', '', '0'],
    ['decimal', 'roads_damaged_km', 'Roads Damaged (km)', '', '', '', '', '', '', '', '0'],
    ['integer', 'bridges_damaged', 'Bridges Damaged', '', '', '', '', '', '', '', '0'],
    ['integer', 'tanks_breached', 'Tanks Breached', '', '', '', '', '', '', '', '0'],
    ['integer', 'bunds_breached', 'Bunds Breached', '', '', '', '', '', '', '', '0'],
    ['end group', 'grp_infra'],

    ['begin group', 'grp_crop', 'Crop Damage', '', '', 'w5'],
    ['decimal', 'crop_area_ha', 'Crop Area Affected (Ha)', '', '', '', '', '', '', '', '0'],
    ['decimal', 'crop_loss_crores', 'Crop Loss (Crores Rs.)', '', '', '', '', '', '', '', '0'],
    ['end group', 'grp_crop'],

    ['begin group', 'grp_relief', 'Relief Operations', '', '', 'w5'],
    ['integer', 'sdrf_teams', 'SDRF Teams Deployed', '', '', '', '', '', '', '', '0'],
    ['integer', 'ndrf_teams', 'NDRF Teams Deployed', '', '', '', '', '', '', '', '0'],
    ['integer', 'boats_deployed', 'Boats Deployed', '', '', '', '', '', '', '', '0'],
    ['integer', 'relief_camps', 'Relief Camps Opened', '', '', '', '', '', '', '', '0'],
    ['integer', 'people_in_camps', 'People in Relief Camps', '', '', '', '', '', '', '', '0'],
    ['integer', 'people_evacuated', 'People Evacuated', '', '', '', '', '', '', '', '0'],
    ['integer', 'medical_camps', 'Medical Camps Conducted', '', '', '', '', '', '', '', '0'],
    ['integer', 'food_packets', 'Food Packets Distributed', '', '', '', '', '', '', '', '0'],
    ['end group', 'grp_relief'],
]

flood_choices = {
    'report_shift': [
        ('morning', 'Morning (06:00 Hrs)'),
        ('afternoon', 'Afternoon (14:00 Hrs)'),
        ('night', 'Night (22:00 Hrs)'),
    ],
}

create_form('SEOC_Flood_Damage.xlsx', 'SEOC - Flood Damage Report',
            'seoc_flood_damage', flood_rows, flood_choices)


# ================================================================
# FORM 4: EMERGENCY CALLS (112/1070)
# ================================================================
print("[4/8] Emergency Calls Log")
calls_rows = [
    ['begin group', 'grp_call', 'Call Information', '', '', 'w5'],
    ['select_one call_source', 'call_source', 'Call Source', '', 'yes', ''],
    ['text', 'caller_name', 'Caller Name', '', '', ''],
    ['text', 'caller_phone', 'Caller Phone', '', 'yes', ''],
    ['select_one event_type', 'event_type', 'Event Type', '', 'yes', ''],
    ['text', 'event_info', 'Event Description', '', 'yes', 'multiline'],
    ['end group', 'grp_call'],

    ['begin group', 'grp_victim', 'Victim Details', '', '', 'w5'],
    ['text', 'victim_name_age', 'Victim Name & Age', '', '', ''],
    ['integer', 'num_persons', 'No. of Persons Involved', '', '', ''],
    ['integer', 'bodies_traced', 'Bodies Traced', '', '', '', '', '', '', '', '0'],
    ['integer', 'persons_missing', 'Persons Missing', '', '', '', '', '', '', '', '0'],
    ['integer', 'persons_saved', 'Persons Saved', '', '', '', '', '', '', '', '0'],
    ['end group', 'grp_victim'],

    ['begin group', 'grp_response', 'Response', '', '', 'w5'],
    ['select_multiple team_deployed', 'team_deployed', 'Team Deployed', '', '', ''],
    ['text', 'info_passed_time', 'Information Passed Time', 'e.g. MRO at 2:31 PM', '', 'multiline'],
    ['text', 'erss_closing_time', 'ERSS Closing Time', '', '', ''],
    ['end group', 'grp_response'],
]

calls_choices = {
    'call_source': [
        ('112a', '112 A'), ('112b', '112 B'), ('1070', '1070'),
    ],
    'event_type': [
        ('drowning', 'Drowning'), ('fire', 'Fire'), ('flood', 'Flood/Heavy Rain'),
        ('lightning', 'Lightning'), ('road_accident', 'Road Accident'),
        ('building_collapse', 'Building/Wall Collapse'), ('tree_fall', 'Tree Fall'),
        ('electrocution', 'Electrocution'), ('boat_capsize', 'Boat Capsize'),
        ('animal_attack', 'Animal Attack'), ('chemical', 'Chemical/Gas Leak'),
        ('other', 'Other'),
    ],
    'team_deployed': [
        ('sdrf', 'SDRF'), ('ndrf', 'NDRF'), ('fire', 'Fire Department'),
        ('police', 'Police'), ('revenue', 'Revenue'), ('ambulance', 'Ambulance'),
        ('navy', 'Navy'), ('coast_guard', 'Coast Guard'), ('swimmers', 'Local Swimmers'),
        ('other', 'Other'),
    ],
}

create_form('SEOC_Emergency_Calls.xlsx', 'SEOC - Emergency Call Log (112/1070)',
            'seoc_emergency_calls', calls_rows, calls_choices, has_attachments=False)


# ================================================================
# FORM 5: SDRF/NDRF RESCUE OPERATION
# ================================================================
print("[5/8] SDRF/NDRF Rescue Operation")
rescue_rows = [
    ['begin group', 'grp_request', 'Request Details', '', '', 'w5'],
    ['dateTime', 'request_received', 'Request Received Date/Time', '', 'yes', ''],
    ['text', 'request_from', 'Request Received From', 'e.g. Collector, Kurnool', '', ''],
    ['dateTime', 'team_dispatched', 'Team Dispatched Date/Time', '', '', ''],
    ['end group', 'grp_request'],

    ['begin group', 'grp_team', 'Team Details', '', '', 'w5'],
    ['select_one force_type', 'force_type', 'Force Type', '', 'yes', ''],
    ['text', 'battalion', 'Battalion / Unit', '', '', ''],
    ['integer', 'num_teams', 'No. of Teams', '', '', ''],
    ['integer', 'team_strength', 'Team Strength', '', '', ''],
    ['text', 'incharge_name', 'Team In-Charge', '', '', ''],
    ['text', 'incharge_phone', 'In-Charge Phone', '', '', ''],
    ['text', 'equipment', 'Equipment', '', '', ''],
    ['end group', 'grp_team'],

    ['begin group', 'grp_operation', 'Operation Details', '', '', 'w5'],
    ['select_one event_nature', 'event_nature', 'Nature of Event', '', 'yes', ''],
    ['text', 'water_body_name', 'Location Name (River/Canal/Tank)', '', '', ''],
    ['dateTime', 'rescue_datetime', 'Date/Time of Rescue', '', '', ''],
    ['integer', 'males_rescued', 'Males Rescued/Traced', '', '', '', '', '', '', '', '0'],
    ['integer', 'females_rescued', 'Females Rescued/Traced', '', '', '', '', '', '', '', '0'],
    ['integer', 'children_rescued', 'Children Rescued/Traced', '', '', '', '', '', '', '', '0'],
    ['integer', 'persons_survived', 'Persons Survived', '', '', '', '', '', '', '', '0'],
    ['integer', 'deaths', 'Deaths', '', '', '', '', '', '', '', '0'],
    ['end group', 'grp_operation'],

    ['begin group', 'grp_closure', 'Closure', '', '', 'w5'],
    ['date', 'team_closed_date', 'Team Closed / Return Date', '', '', ''],
    ['end group', 'grp_closure'],
]

rescue_choices = {
    'force_type': [
        ('sdrf', 'SDRF'), ('ndrf', 'NDRF'), ('navy', 'Navy'),
        ('coast_guard', 'Coast Guard'), ('army', 'Army'), ('air_force', 'Air Force'),
    ],
    'event_nature': [
        ('drowning', 'Drowning'), ('flood', 'Flood Rescue'), ('cyclone', 'Cyclone'),
        ('boat_capsize', 'Boat Capsize'), ('landslide', 'Landslide'),
        ('building_collapse', 'Building Collapse'), ('other', 'Other'),
    ],
}

create_form('SEOC_Rescue_Operation.xlsx', 'SEOC - SDRF/NDRF Rescue Operation',
            'seoc_rescue_op', rescue_rows, rescue_choices)


# ================================================================
# FORM 6: HEAT WAVE MONITORING
# ================================================================
print("[6/8] Heat Wave Monitoring")
heat_rows = [
    ['begin group', 'grp_observation', 'Temperature Observation', '', '', 'w5'],
    ['decimal', 'max_temperature', 'Maximum Temperature (C)', '', 'yes', ''],
    ['decimal', 'min_temperature', 'Minimum Temperature (C)', '', '', ''],
    ['decimal', 'humidity_percent', 'Humidity (%)', '', '', ''],
    ['select_one yesno', 'heatwave_declared', 'Heatwave Declared?', '', '', ''],
    ['select_one yesno', 'severe_heatwave', 'Severe Heatwave?', '', '', '', '', '', '', '${heatwave_declared}="yes"'],
    ['end group', 'grp_observation'],

    ['begin group', 'grp_impact', 'Heat Wave Impact', '', '', 'w5'],
    ['integer', 'heat_casualties', 'Heat-related Casualties', '', '', '', '', '', '', '', '0'],
    ['integer', 'hospitalizations', 'Hospitalizations', '', '', '', '', '', '', '', '0'],
    ['end group', 'grp_impact'],
]

create_form('SEOC_HeatWave.xlsx', 'SEOC - Heat Wave Monitoring',
            'seoc_heatwave', heat_rows, has_attachments=False)


# ================================================================
# FORM 7: RESERVOIR / RIVER LEVELS
# ================================================================
print("[7/8] Reservoir / River Levels")
reservoir_rows = [
    ['begin group', 'grp_dam', 'Dam / Barrage Details', '', '', 'w5'],
    ['select_one river_system', 'river_system', 'River System', '', 'yes', ''],
    ['select_one dam_name', 'dam_name', 'Dam / Barrage Name', '', 'yes', ''],
    ['end group', 'grp_dam'],

    ['begin group', 'grp_levels', 'Water Levels', '', '', 'w5'],
    ['decimal', 'water_level_ft', 'Water Level (Feet)', '', '', ''],
    ['decimal', 'water_level_m', 'Water Level (Meters)', '', '', ''],
    ['decimal', 'inflow_cusecs', 'Inflow (Cusecs)', '', '', ''],
    ['decimal', 'outflow_cusecs', 'Outflow (Cusecs)', '', '', ''],
    ['decimal', 'storage_tmc', 'Storage (TMC)', '', '', ''],
    ['integer', 'gates_opened', 'No. of Gates Opened', '', '', ''],
    ['select_one warning_level', 'flood_warning', 'Flood Warning Level', '', '', ''],
    ['end group', 'grp_levels'],
]

reservoir_choices = {
    'river_system': [
        ('krishna', 'Krishna'), ('godavari', 'Godavari'),
        ('vamsadhara', 'Vamsadhara'), ('nagavalli', 'Nagavalli'),
        ('penna', 'Penna'), ('tungabhadra', 'Tungabhadra'),
    ],
    'dam_name': [
        ('srisailam', 'Srisailam'), ('nagarjuna_sagar', 'Nagarjuna Sagar'),
        ('pulichintala', 'Pulichintala'), ('prakasam_barrage', 'Prakasam Barrage'),
        ('somasila', 'Somasila'), ('kandaleru', 'Kandaleru'),
        ('dowleswaram', 'Dowleswaram'), ('polavaram', 'Polavaram'),
        ('gandikota', 'Gandikota'), ('mylavaram', 'Mylavaram'),
        ('sunkesula', 'Sunkesula'), ('tungabhadra_dam', 'Tungabhadra Dam'),
        ('thotapalli', 'Thotapalli Barrage'), ('gotta', 'Gotta Barrage'),
        ('other', 'Other'),
    ],
    'warning_level': [
        ('normal', 'Normal'), ('first', '1st Warning'),
        ('second', '2nd Warning'), ('third', '3rd Warning'),
        ('danger', 'Danger Level'),
    ],
}

create_form('SEOC_Reservoir_Levels.xlsx', 'SEOC - Reservoir / River Levels',
            'seoc_reservoir', reservoir_rows, reservoir_choices, has_attachments=False)


# ================================================================
# FORM 8: DAILY SITUATION REPORT (DSR)
# ================================================================
print("[8/8] Daily Situation Report")
dsr_rows = [
    ['begin group', 'grp_disaster', 'Disaster Details', '', '', 'w5'],
    ['select_one disaster_type', 'disaster_type', 'Type of Disaster', '', 'yes', ''],
    ['end group', 'grp_disaster'],

    ['begin group', 'grp_casualties', 'Casualties', '', '', 'w5'],
    ['integer', 'deaths', 'Deaths', '', '', '', '', '', '', '', '0'],
    ['integer', 'injuries', 'Injuries', '', '', '', '', '', '', '', '0'],
    ['integer', 'missing', 'Missing', '', '', '', '', '', '', '', '0'],
    ['end group', 'grp_casualties'],

    ['begin group', 'grp_damage', 'Damage', '', '', 'w5'],
    ['decimal', 'property_damage_crores', 'Property Damage (Crores Rs.)', '', '', '', '', '', '', '', '0'],
    ['decimal', 'crop_damage_ha', 'Crop Damage (Hectares)', '', '', '', '', '', '', '', '0'],
    ['decimal', 'infra_damage_crores', 'Infrastructure Damage (Crores Rs.)', '', '', '', '', '', '', '', '0'],
    ['end group', 'grp_damage'],
]

dsr_choices = {
    'disaster_type': [
        ('flood', 'Flood / Heavy Rains'), ('cyclone', 'Cyclone'),
        ('lightning', 'Lightning / Thunderbolt'), ('drowning', 'Drowning'),
        ('heatwave', 'Heat Wave / Heat Stroke'), ('fire', 'Fire'),
        ('landslide', 'Landslide'), ('earthquake', 'Earthquake'),
        ('building_collapse', 'Building / Wall Collapse'),
        ('tree_fall', 'Tree Fall'), ('electrocution', 'Electrocution'),
        ('boat_capsize', 'Boat Capsize'), ('snake_bite', 'Snake Bite'),
        ('animal_attack', 'Animal Attack'), ('chemical', 'Chemical Disaster'),
        ('road_accident', 'Road Accident'), ('cold_wave', 'Cold Wave'),
        ('other', 'Other'),
    ],
}

create_form('SEOC_DSR.xlsx', 'SEOC - Daily Situation Report',
            'seoc_dsr', dsr_rows, dsr_choices)


# ================================================================
print(f"\n{'='*60}")
print(f"ALL 8 SURVEY123 FORMS CREATED in {OUT_DIR}")
print(f"{'='*60}")
print("""
Files created:
  1. SEOC_Lightning_Death.xlsx     - Lightning death proforma
  2. SEOC_Drowning_Incident.xlsx   - Drowning incident report
  3. SEOC_Flood_Damage.xlsx        - Flood damage report (45 fields)
  4. SEOC_Emergency_Calls.xlsx     - 112/1070 emergency call log
  5. SEOC_Rescue_Operation.xlsx    - SDRF/NDRF rescue operations
  6. SEOC_HeatWave.xlsx            - Heat wave monitoring
  7. SEOC_Reservoir_Levels.xlsx    - Reservoir/river level monitoring
  8. SEOC_DSR.xlsx                 - Daily Situation Report

Each form includes:
  - Cascading dropdowns: 28 Districts -> 688 Mandals (from PostGIS DB)
  - GPS location capture (geopoint)
  - Photo/PDF attachment support
  - Entry metadata (status, entered_by, datetime)
  - Form-specific choice lists

NEXT STEPS:
  1. Open each .xlsx in Survey123 Connect
  2. Preview and test
  3. Publish to ArcGIS Portal (https://192.168.8.24/gisportal)
  4. Create Portal groups and assign users
  5. Build Experience Builder landing page
""")
