#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""SEOC Staff Registration — Survey123 XLSForm Generator"""

import sys, os
os.environ['PYTHONIOENCODING'] = 'utf-8'
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

OUT = 'h:/SEOC_DB_PORTAL/SEOC_PORTAL/SEOC_Staff_Registration.xlsx'
os.makedirs(os.path.dirname(OUT), exist_ok=True)

wb = openpyxl.Workbook()

# ── Styles ──
HDR_FILL = PatternFill(start_color='1a365d', end_color='1a365d', fill_type='solid')
HDR_FONT = Font(bold=True, color='FFFFFF', size=11)
GRP_FILL = PatternFill(start_color='ebf4ff', end_color='ebf4ff', fill_type='solid')
GRP_FONT = Font(bold=True, color='1a365d', size=11)

def style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')

def add_row(ws, row, is_group=False):
    ws.append(row)
    if is_group:
        for cell in ws[ws.max_row]:
            cell.fill = GRP_FILL
            cell.font = GRP_FONT

# ════════════════════════════════════════════
# SURVEY SHEET
# ════════════════════════════════════════════
ws = wb.active
ws.title = 'survey'
headers = ['type', 'name', 'label', 'hint', 'required', 'appearance',
           'calculation', 'constraint', 'constraint_message', 'relevant',
           'default', 'choice_filter', 'read_only']
ws.append(headers)
style_header(ws)

# ── Metadata (hidden) ──
add_row(ws, ['begin group', 'grp_meta', 'Record Info', '', '', 'w5'], True)
add_row(ws, ['text', 'entry_status', 'Status', '', '', 'hidden', "'Pending'", '', '', '', 'Pending', '', 'yes'])
add_row(ws, ['text', 'registered_by', 'Registered By', '', '', 'hidden', '', '', '', '', '', '', 'yes'])
add_row(ws, ['dateTime', 'registration_date', 'Registration Date', '', '', '', 'now()', '', '', '', '', '', 'yes'])
add_row(ws, ['end group', 'grp_meta'])

# ══════════════════════════════════════
# SECTION 1: PERSONAL DETAILS
# ══════════════════════════════════════
add_row(ws, ['begin group', 'sec_personal', 'Section 1: Personal Details', '', '', 'w5 field-list'], True)

add_row(ws, ['text', 'full_name', 'Full Name', 'As per official records', 'yes', ''])
add_row(ws, ['date', 'date_of_birth', 'Date of Birth', '', '', ''])
add_row(ws, ['select_one gender', 'gender', 'Gender', '', '', ''])
add_row(ws, ['text', 'employee_id', 'Employee ID / Staff No.', 'e.g. APSDMA-001', '', ''])
add_row(ws, ['image', 'photo', 'Passport Photo', 'Max 2MB', '', 'annotate'])

add_row(ws, ['end group', 'sec_personal'])

# ══════════════════════════════════════
# SECTION 2: APSDMA POSTING
# ══════════════════════════════════════
add_row(ws, ['begin group', 'sec_posting', 'Section 2: APSDMA Posting', '', '', 'w5 field-list'], True)

add_row(ws, ['select_one designation', 'designation', 'Designation', '', 'yes', 'autocomplete'])
add_row(ws, ['text', 'designation_other', 'Specify Designation', '', '', '',
             '', '', '', '${designation}="other"'])

add_row(ws, ['select_one section_wing', 'section_wing', 'Section / Wing', '', '', ''])
add_row(ws, ['text', 'section_wing_other', 'Specify Section', '', '', '',
             '', '', '', '${section_wing}="other"'])

add_row(ws, ['select_one posting_type', 'posting_type', 'Posting Type', '', 'yes', ''])
add_row(ws, ['date', 'date_of_joining', 'Date of Joining APSDMA', '', '', ''])
add_row(ws, ['select_one staff_status', 'staff_status', 'Current Status', '', 'yes', '', '', '', '', '', 'active'])
add_row(ws, ['select_one shift_duty', 'shift_duty', 'Shift / Duty Pattern', '', '', ''])

add_row(ws, ['end group', 'sec_posting'])

# ══════════════════════════════════════
# SECTION 3: CONTACT
# ══════════════════════════════════════
add_row(ws, ['begin group', 'sec_contact', 'Section 3: Contact Information', '', '', 'w5 field-list'], True)

add_row(ws, ['text', 'mobile_primary', 'Personal Mobile', '10-digit number', 'yes', '',
             '', 'regex(.,\"^[0-9]{10}$\")', 'Enter valid 10-digit mobile number'])
add_row(ws, ['text', 'mobile_alternate', 'Alternate Mobile', '', '', ''])
add_row(ws, ['text', 'whatsapp_number', 'WhatsApp Number', 'If different from primary', '', ''])
add_row(ws, ['text', 'email_official', 'Email (Official)', 'e.g. name@ap.gov.in', '', ''])
add_row(ws, ['text', 'seoc_intercom', 'SEOC Intercom No.', '', '', ''])

add_row(ws, ['end group', 'sec_contact'])

# ══════════════════════════════════════
# SECTION 4: PORTAL ACCESS
# ══════════════════════════════════════
add_row(ws, ['begin group', 'sec_portal', 'Section 4: Portal Access', '', '', 'w5 field-list'], True)

add_row(ws, ['note', 'portal_note', '**Assign Esri Portal role and disaster modules for this staff member.**', '', '', ''])
add_row(ws, ['select_one portal_role', 'portal_role', 'Portal Role', 'Determines what the user can do', 'yes', ''])
add_row(ws, ['select_multiple disaster_module', 'assigned_modules', 'Assigned Disaster Modules',
             'Which modules can this user access?', 'yes', ''])
add_row(ws, ['text', 'portal_username', 'Portal Username', 'e.g. deo_ramesh, do_suresh', 'yes', ''])
add_row(ws, ['select_one account_status', 'account_created', 'Portal Account Created?', '', '', '', '', '', '', '', 'no'])

add_row(ws, ['end group', 'sec_portal'])

# ══════════════════════════════════════
# SECTION 5: REMARKS
# ══════════════════════════════════════
add_row(ws, ['begin group', 'sec_remarks', 'Section 5: Remarks', '', '', 'w5'], True)

add_row(ws, ['text', 'remarks', 'Remarks / Notes', 'Any additional information', '', 'multiline'])

add_row(ws, ['end group', 'sec_remarks'])

# Set column widths
for i, w in enumerate([18, 22, 40, 30, 8, 15, 20, 25, 25, 25, 12, 25, 8]):
    ws.column_dimensions[chr(65+i)].width = w


# ════════════════════════════════════════════
# CHOICES SHEET
# ════════════════════════════════════════════
ws_ch = wb.create_sheet('choices')
ws_ch.append(['list_name', 'name', 'label'])
style_header(ws_ch)

# Gender
for v, l in [('male', 'Male'), ('female', 'Female')]:
    ws_ch.append(['gender', v, l])

# Designation
designations = [
    ('md', 'Managing Director, APSDMA'),
    ('scs', 'Special Chief Secretary'),
    ('ds', 'Deputy Secretary'),
    ('as', 'Assistant Secretary'),
    ('so', 'Section Officer'),
    ('seoc_incharge', 'SEOC Incharge'),
    ('seoc_duty_officer', 'SEOC Duty Officer'),
    ('deo', 'Data Entry Operator'),
    ('gis_analyst', 'GIS Analyst'),
    ('it_admin', 'IT Administrator'),
    ('system_admin', 'System Administrator'),
    ('constable', 'Constable'),
    ('head_constable', 'Head Constable'),
    ('rsi', 'RSI'),
    ('asi', 'ASI'),
    ('si', 'SI'),
    ('consultant', 'Consultant'),
    ('programmer', 'Programmer / Developer'),
    ('driver', 'Driver'),
    ('attender', 'Attender'),
    ('mts', 'MTS (Multi Tasking Staff)'),
    ('other', 'Other (specify below)'),
]
for v, l in designations:
    ws_ch.append(['designation', v, l])

# Section / Wing
sections = [
    ('seoc_control', 'SEOC Control Room'),
    ('gis_cell', 'GIS Cell'),
    ('it_cell', 'IT Cell'),
    ('administration', 'Administration'),
    ('finance', 'Finance & Accounts'),
    ('operations', 'Operations'),
    ('planning', 'Planning & Preparedness'),
    ('early_warning', 'Early Warning'),
    ('capacity_building', 'Capacity Building & Training'),
    ('response', 'Response & Relief'),
    ('mitigation', 'Mitigation'),
    ('other', 'Other (specify below)'),
]
for v, l in sections:
    ws_ch.append(['section_wing', v, l])

# Posting Type
for v, l in [('regular', 'Regular'), ('deputation', 'Deputation'),
             ('contract', 'Contract'), ('outsourcing', 'Outsourcing'),
             ('attachment', 'Attachment')]:
    ws_ch.append(['posting_type', v, l])

# Status
for v, l in [('active', 'Active'), ('on_leave', 'On Leave'),
             ('transferred', 'Transferred'), ('relieved', 'Relieved'),
             ('suspended', 'Suspended')]:
    ws_ch.append(['staff_status', v, l])

# Shift
for v, l in [('day', 'Day Shift (8AM-8PM)'), ('night', 'Night Shift (8PM-8AM)'),
             ('general', 'General (9:30AM-6PM)'), ('rotational', 'Rotational (24x7)'),
             ('on_call', 'On Call')]:
    ws_ch.append(['shift_duty', v, l])

# Portal Role
portal_roles = [
    ('admin', 'Admin (full control, manage users & forms)'),
    ('seoc_incharge', 'SEOC Incharge (approve, view all, dashboards)'),
    ('duty_officer', 'Duty Officer (ratify/approve DEO entries)'),
    ('deo', 'DEO - Data Entry Operator (enter data only)'),
    ('viewer', 'Viewer (read-only, dashboards only)'),
]
for v, l in portal_roles:
    ws_ch.append(['portal_role', v, l])

# Disaster Modules
modules = [
    ('lightning', 'Lightning Deaths'),
    ('drowning', 'Drowning Incidents'),
    ('floods', 'Flood Damage Reports'),
    ('emergency_calls', 'Emergency Calls (112/1070)'),
    ('rescue_ops', 'SDRF/NDRF Rescue Operations'),
    ('heatwave', 'Heat Wave Monitoring'),
    ('reservoir', 'Reservoir / River Levels'),
    ('dsr', 'Daily Situation Report'),
    ('all', 'All Modules'),
]
for v, l in modules:
    ws_ch.append(['disaster_module', v, l])

# Account status
for v, l in [('yes', 'Yes - Created'), ('no', 'No - Pending'), ('disabled', 'Disabled')]:
    ws_ch.append(['account_status', v, l])

# Column widths
ws_ch.column_dimensions['A'].width = 18
ws_ch.column_dimensions['B'].width = 20
ws_ch.column_dimensions['C'].width = 45


# ════════════════════════════════════════════
# SETTINGS SHEET
# ════════════════════════════════════════════
ws_set = wb.create_sheet('settings')
ws_set.append(['form_title', 'form_id', 'version', 'style'])
style_header(ws_set)
ws_set.append(['APSDMA Staff Registration', 'apsdma_staff_registration', '2026.1', 'theme-grid'])
ws_set.column_dimensions['A'].width = 30
ws_set.column_dimensions['B'].width = 30


# ── Save ──
wb.save(OUT)
print(f"Created: {OUT}")
print(f"Sheets: survey ({ws.max_row} rows), choices ({ws_ch.max_row} rows), settings")
print("\nForm structure:")
print("  Section 1: Personal Details (5 fields)")
print("  Section 2: APSDMA Posting (8 fields)")
print("  Section 3: Contact Information (5 fields)")
print("  Section 4: Portal Access (5 fields)")
print("  Section 5: Remarks (1 field)")
print(f"\nDropdown options:")
print(f"  Designations: {len(designations)}")
print(f"  Sections/Wings: {len(sections)}")
print(f"  Portal Roles: {len(portal_roles)}")
print(f"  Disaster Modules: {len(modules)}")
