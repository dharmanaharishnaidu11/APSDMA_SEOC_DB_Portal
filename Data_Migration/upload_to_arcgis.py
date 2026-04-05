#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Upload existing data to ArcGIS Feature Services
Sources:
  1. all_deaths_extracted.csv → SEOC_Lightning_Deaths (648 records)
  2. PostgreSQL seoc.emergency_calls → SEOC_Emergency_Calls (161 records)
  3. All Calls 2026.xlsx → SEOC_Emergency_Calls (additional 2026 data)
"""

import sys, os, csv, json, time, warnings
os.environ['PYTHONIOENCODING'] = 'utf-8'
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')
warnings.filterwarnings('ignore')

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from _config_loader import load_config

import requests, urllib3
urllib3.disable_warnings()

_cfg = load_config()
PORTAL = _cfg['arcgis_portal']['url']
SERVER = PORTAL.replace('/gisportal', '/gisserver') + '/rest/services/Hosted'

# Get token
r = requests.post(f'{PORTAL}/sharing/rest/generateToken',
    data={'username': _cfg['arcgis_portal']['username'],
          'password': _cfg['arcgis_portal']['password'],
          'client':'referer','referer':'https://apsdmagis.ap.gov.in','f':'json'},
    verify=False)
TOKEN = r.json()['token']
print(f"Token acquired\n")


def add_features_batch(service_name, features, batch_size=50):
    """Upload features in batches to avoid timeout"""
    url = f'{SERVER}/{service_name}/FeatureServer/0/addFeatures'
    total = len(features)
    added = 0
    errors = 0

    for i in range(0, total, batch_size):
        batch = features[i:i+batch_size]
        r = requests.post(url,
            data={'features': json.dumps(batch), 'f': 'json', 'token': TOKEN},
            verify=False, timeout=120)
        result = r.json()
        results = result.get('addResults', [])
        batch_ok = sum(1 for rr in results if rr.get('success'))
        batch_err = sum(1 for rr in results if not rr.get('success'))
        added += batch_ok
        errors += batch_err

        if batch_err > 0:
            for rr in results:
                if not rr.get('success'):
                    print(f"    Error: {rr.get('error', {}).get('description', 'unknown')[:80]}")

        pct = round((i + len(batch)) / total * 100)
        print(f"    Batch {i//batch_size + 1}: {batch_ok}/{len(batch)} OK ({pct}% done)")
        time.sleep(0.5)

    return added, errors


def parse_date(date_str):
    """Parse various date formats to epoch ms"""
    if not date_str:
        return None
    from datetime import datetime
    formats = ['%m-%d-%Y', '%Y-%m-%d', '%m/%d/%Y', '%d-%m-%Y', '%Y-%m-%d %H:%M:%S']
    for fmt in formats:
        try:
            dt = datetime.strptime(str(date_str).strip(), fmt)
            return int(dt.timestamp() * 1000)
        except:
            continue
    return None


# ══════════════════════════════════════════════════════════
# 1. LIGHTNING DEATHS from CSV
# ══════════════════════════════════════════════════════════
print("=" * 60)
print("1. LIGHTNING DEATHS (all_deaths_extracted.csv → ArcGIS)")
print("=" * 60)

csv_path = os.path.join(os.path.dirname(__file__), 'all_deaths_extracted.csv')
coords_path = os.path.join(os.path.dirname(__file__), 'filled_coords_all.csv')

# Load geocoded coordinates
coords = {}
with open(coords_path, 'r', encoding='utf-8') as f:
    reader = csv.DictReader(f)
    for row in reader:
        sno = row.get('SNo', '').strip()
        lat = row.get('Filled_Lat', '').strip()
        lon = row.get('Filled_Long', '').strip()
        if sno and lat and lon:
            try:
                coords[sno] = (float(lat), float(lon))
            except:
                pass
print(f"  Loaded {len(coords)} geocoded coordinates")

features = []
with open(csv_path, 'r', encoding='utf-8') as f:
    reader = csv.DictReader(f)
    for row in reader:
        sno = row.get('SNo', '').strip().replace('.0', '')
        lat = None
        lon = None

        # Try direct coords from CSV
        try:
            lat = float(row.get('Lat', '').strip())
            lon = float(row.get('Long', '').strip())
        except:
            pass

        # Fallback to filled_coords
        if (not lat or not lon) and sno in coords:
            lat, lon = coords[sno]

        attrs = {
            'incident_date': parse_date(row.get('Date', '')),
            'victim_name': row.get('Name', '').strip() or None,
            'aadhar_number': row.get('Aadhar', '').strip() or None,
            'gender': row.get('Gender', '').strip() or None,
            'occupation': row.get('Occupation', '').strip() or None,
            'district': row.get('District', '').strip() or None,
            'mandal': row.get('Mandal', '').strip() or None,
            'village': row.get('Village', '').strip() or None,
            'habitation': row.get('Habitation', '').strip() or None,
            'death_location': row.get('Location_Type', '').strip() or None,
            'economic_status': row.get('Economic_Status', '').strip() or None,
            'kin_name': row.get('Next_Kin', '').strip() or None,
            'entry_status': 'Approved',
            'entered_by': 'data_migration',
            'approved_by': 'data_migration',
        }

        # Parse age
        age = row.get('Age', '').strip().replace('.0', '')
        if age:
            try:
                attrs['age'] = int(float(age))
            except:
                pass

        # Remove None values
        attrs = {k: v for k, v in attrs.items() if v is not None}

        feat = {'attributes': attrs}
        if lat and lon and abs(lat) <= 90 and abs(lon) <= 180:
            feat['geometry'] = {'x': lon, 'y': lat, 'spatialReference': {'wkid': 4326}}

        features.append(feat)

print(f"  Parsed {len(features)} records from CSV")
with_geom = sum(1 for f in features if 'geometry' in f)
print(f"  With coordinates: {with_geom}")

added, errors = add_features_batch('SEOC_Lightning_Deaths', features)
print(f"\n  Result: {added} added, {errors} errors\n")


# ══════════════════════════════════════════════════════════
# 2. EMERGENCY CALLS from PostgreSQL
# ══════════════════════════════════════════════════════════
print("=" * 60)
print("2. EMERGENCY CALLS (PostgreSQL → ArcGIS)")
print("=" * 60)

import psycopg2

conn = psycopg2.connect(**load_config()['database'])
cur = conn.cursor()

cur.execute("""
    SELECT call_date, call_time, call_source, caller_name, caller_phone,
           event_type, event_info, district, mandal, village,
           victim_name_age, num_persons, bodies_traced, persons_missing,
           persons_saved, team_deployed, entry_status, entered_by
    FROM seoc.emergency_calls ORDER BY id
""")

columns = [d[0] for d in cur.description]
rows = cur.fetchall()
conn.close()

features = []
for row in rows:
    attrs = {}
    for i, col in enumerate(columns):
        val = row[i]
        if val is None:
            continue
        if col == 'call_date' and val:
            from datetime import datetime, date
            import calendar
            if isinstance(val, (datetime, date)):
                try:
                    attrs[col] = int(calendar.timegm(val.timetuple())) * 1000
                except:
                    pass
            continue
        if isinstance(val, int):
            attrs[col] = val
        else:
            attrs[col] = str(val)

    if not attrs.get('entry_status'):
        attrs['entry_status'] = 'Approved'
    if not attrs.get('entered_by'):
        attrs['entered_by'] = 'data_migration'

    features.append({'attributes': attrs})

print(f"  Parsed {len(features)} records from PostgreSQL")
added, errors = add_features_batch('SEOC_Emergency_Calls', features)
print(f"\n  Result: {added} added, {errors} errors\n")


# ══════════════════════════════════════════════════════════
# 3. ALL CALLS 2026 from Excel (Jan sheet)
# ══════════════════════════════════════════════════════════
print("=" * 60)
print("3. ALL CALLS 2026 (Excel → ArcGIS)")
print("=" * 60)

try:
    import openpyxl
    xlsx_path = "S:/SEOC_PORTAL/H_Drive_Backup/SEOC_DB_PORTAL/SEOC_COMMUNICATIONS/All Calls (112_1070) - 2026.xlsx"
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    # Use the consolidated sheet
    sheet_name = '112A  112B   1070'
    if sheet_name not in wb.sheetnames:
        sheet_name = wb.sheetnames[0]

    ws = wb[sheet_name]
    rows_data = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows_data) > 1:
        headers = rows_data[0]
        print(f"  Sheet: {sheet_name}, Headers: {[str(h)[:25] for h in headers[:12]]}")

        features_xl = []
        for row in rows_data[1:]:
            if not row or not any(row):
                continue

            attrs = {
                'entry_status': 'Approved',
                'entered_by': 'excel_migration',
                'approved_by': 'data_migration'
            }

            # Map columns
            for i, val in enumerate(row):
                if val is None or str(val).strip() == '':
                    continue
                val = str(val).strip() if not isinstance(val, (int, float)) else val

                if i == 1:  # call_source
                    src = str(val).strip()
                    if '112' in src.lower():
                        attrs['call_source'] = '112a' if 'a' in src.lower() else '112b'
                    elif '1070' in src:
                        attrs['call_source'] = '1070'
                    else:
                        attrs['call_source'] = src[:10]
                elif i == 2:  # date
                    from datetime import datetime
                    if isinstance(val, datetime):
                        import calendar
                        attrs['call_date'] = int(calendar.timegm(val.timetuple())) * 1000
                    else:
                        d = parse_date(str(val))
                        if d:
                            attrs['call_date'] = d
                elif i == 4:  # time
                    attrs['call_time'] = str(val)[:20]
                elif i == 5:  # caller
                    attrs['caller_name'] = str(val)[:200]
                elif i == 6:  # description
                    attrs['event_info'] = str(val)[:2000]
                elif i == 7:  # team deployed
                    attrs['team_deployed'] = str(val)[:500]
                elif i == 8:  # event type
                    attrs['event_type'] = str(val).lower()[:50]
                elif i == 9:  # district
                    attrs['district'] = str(val)[:100]
                elif i == 10:  # mandal
                    attrs['mandal'] = str(val)[:100]

            if attrs.get('district') or attrs.get('event_info'):
                features_xl.append({'attributes': attrs})

        print(f"  Parsed {len(features_xl)} records from Excel")
        if features_xl:
            added, errors = add_features_batch('SEOC_Emergency_Calls', features_xl)
            print(f"\n  Result: {added} added, {errors} errors")
    else:
        print("  No data rows found")

except Exception as e:
    print(f"  Error processing Excel: {e}")


# ══════════════════════════════════════════════════════════
# FINAL VERIFICATION
# ══════════════════════════════════════════════════════════
print("\n" + "=" * 60)
print("VERIFICATION")
print("=" * 60)

services = ['SEOC_Lightning_Deaths', 'SEOC_Emergency_Calls', 'SEOC_Drowning_Incidents',
            'SEOC_Flood_Damage', 'SEOC_Rescue_Operations', 'SEOC_HeatWave',
            'SEOC_Reservoir_Levels', 'SEOC_DSR', 'SEOC_Staff_Registration']

for svc in services:
    r = requests.get(f'{SERVER}/{svc}/FeatureServer/0/query',
        params={'where': '1=1', 'returnCountOnly': 'true', 'f': 'json', 'token': TOKEN},
        verify=False)
    count = r.json().get('count', '?')
    print(f"  {svc:<30} {count} records")

print("\nDone!")
